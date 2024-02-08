import os

import openpyxl
import pandas as pd

from openpyxl.styles import PatternFill, Alignment, Font
from django.conf import settings



class ReportGeneration:
    """Generating report"""
    

    def __init__(self, filepath: str) -> None:
        wb = openpyxl.load_workbook(filename=filepath, read_only=True)
        self.input_file_path = filepath
        self.time_create = filepath.split('input_data_')[-1].replace('.xlsx', '')
        self.ws = wb.active
        self.data_for_write = []
        self.wb_report = openpyxl.Workbook()
        self.ws_report = self.wb_report.active
        self.ws_report.title = 'Отчет'


    def __create_report_cap(self):
        # add headers
        self.ws_report.append(['Филиал', 'Сотрудник', 'Налоговая база', 'Налог', 'Отклонения'])
        self.ws_report.append([None, None, None, 'Исчислено всего', 'Исчислено всего по формуле'])

        # merging cells
        for interval in ('A1:A2', 'B1:B2', 'C1:C2', 'D1:E1', 'F1:F2'):
            self.ws_report.merge_cells(interval)

        # add styles
        center_align = Alignment(horizontal='center', vertical='center')
        light_blue = PatternFill(fill_type='solid',
                                 start_color='cbe4e5',
                                 end_color='cbe4e5')
        font = Font(bold=True, color='00005c')
        
        for col in 'ABCDEF':
            cell1 = self.ws_report[col + '1']
            cell2 = self.ws_report[col + '2']

            for cell in (cell1, cell2):
                cell.fill = light_blue
                cell.alignment = center_align
                cell.font = font
        

    def __generate_report_data(self):
        for row in self.ws.iter_rows(values_only=True):
            # original columns
            branch = row[0]
            employee = row[1]
            
            try:
                tax_base = round(float(row[4]), 2)
                total_calculated = round(float(row[5]), 2)
            except Exception:
                continue

            if None in (branch, employee, tax_base, total_calculated):   # if no data
                continue

            # new columns
            if tax_base <= 5_000_000:
                total_calculated_by_formula = tax_base * 0.13
            else:
                total_calculated_by_formula = 5_000_000 * 0.13 + (tax_base - 5_000_000) * 0.15

            total_calculated_by_formula = round(total_calculated_by_formula, 2)

            deviations= round(total_calculated - total_calculated_by_formula, 2)
            # deviations = difference if difference != 0 else None

            # append data for writing
            self.data_for_write.append([branch, employee,
                                   tax_base, total_calculated,
                                   total_calculated_by_formula,
                                   deviations])


    def __colorize_deviations_col(self):
        red_fill = PatternFill(fill_type="solid",
                               start_color="00FF0000",
                               end_color="00FF0000")
        green_fill = PatternFill(start_color="0000FF00",
                                 end_color="0000FF00",
                                 fill_type="solid")
        
        for row in self.ws_report.iter_rows(min_row=3):
            cell = row[5]
            cell.fill = green_fill if cell.value == 0 else red_fill


    def __sort_data(self):
        df = pd.DataFrame(data=self.data_for_write,
                          columns=['f', 'e', 'tb', 'tc', 'tcb', 'd'])
        df = df.sort_values('d', ascending=False)

        return df.values.tolist()

    
    def __record_data_to_excel(self, sorted_data: list) -> None:
        for row in sorted_data:
            self.ws_report.append(row)


    def __set_columns_width(self):
        for col in self.ws_report.iter_cols():
            max_width = 0
            for cell in col:
                if not cell.value: continue

                width = len(str(cell.value))
                if width > max_width:
                    max_width = width + 10
                    self.ws_report.column_dimensions[cell.column_letter].width = max_width


    def create_report(self) -> str:
        self.__create_report_cap()
        self.__generate_report_data()
        sort_data = self.__sort_data()
        self.__record_data_to_excel(sort_data)
        self.__colorize_deviations_col()
        self.__set_columns_width()

        path = os.path.join(settings.MEDIA_ROOT, f'report_{self.time_create}.xlsx')
        self.wb_report.save(path)

        return path


    def __del__(self):
        os.remove(self.input_file_path)